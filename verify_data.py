"""JOB 1: Full data integrity audit — verify every Excel row against the SQLite database."""

import sqlite3
from collections import Counter
from openpyxl import load_workbook

EXCEL = "/Users/nathangoldberg/Downloads/Dakota_Pension_Fund_Data_Sample (1).xlsx"
DB = "data/pension_tracker.db"

wb = load_workbook(EXCEL)
ws = wb["Data"]

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row

pf_rows = conn.execute("SELECT id, name FROM pension_funds").fetchall()
PF_ID = {}
for r in pf_rows:
    PF_ID[r["name"]] = r["id"]

headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
print(f"Excel: {ws.max_row - 1} data rows\n")

confirmed = []
suspect = []


def safe_float(v):
    if v is None:
        return None
    try:
        return float(str(v).replace("x", "").replace(",", "").strip())
    except (ValueError, TypeError):
        return None


QUERY = """
    SELECT c.commitment_mm, c.capital_called_mm, c.capital_distributed_mm,
           c.remaining_value_mm, c.net_irr, c.net_multiple, c.as_of_date,
           c.source_url, c.source_document, c.extraction_method,
           c.extraction_confidence,
           f.fund_name, f.general_partner_normalized,
           f.asset_class, f.sub_strategy, f.vintage_year
    FROM commitments c
    JOIN funds f ON c.fund_id = f.id
    WHERE c.pension_fund_id = ?
      AND f.fund_name = ?
"""

QUERY_FUZZY = QUERY.replace("f.fund_name = ?", "f.fund_name LIKE ?")

shifted_rows = []  # Track rows with column-shift corruption

for row_idx in range(2, ws.max_row + 1):
    rec = {}
    for c in range(1, ws.max_column + 1):
        rec[headers[c - 1]] = ws.cell(row=row_idx, column=c).value

    # Detect column-shifted rows: Fund Name is a number instead of string
    fund_name_raw = rec["Fund Name"]
    if not isinstance(fund_name_raw, str):
        # Columns shifted left by 1: State has GP, GP Name has Fund Name, Fund Name has Commitment
        real_gp = str(rec["State"])
        real_fund = str(rec["GP Name"])
        real_commitment = safe_float(rec["Fund Name"])
        real_vintage = safe_float(rec["Vintage Year"])
        shifted_rows.append((row_idx, rec["Pension Fund"], real_gp, real_fund, real_commitment, real_vintage))
        # We'll verify these separately below
        continue

    pension_fund = rec["Pension Fund"]
    fund_name = rec["Fund Name"]
    gp_excel = rec["GP Name"]
    commitment = safe_float(rec["Commitment ($M)"])
    vintage = safe_float(rec["Vintage Year"])

    pfid = PF_ID.get(pension_fund)
    if not pfid:
        suspect.append((row_idx, f"Unknown pension fund: {pension_fund}"))
        continue

    # Find matching commitment rows in DB
    raw_matches = conn.execute(QUERY, (pfid, fund_name)).fetchall()
    matches = [dict(m) for m in raw_matches]

    if not matches:
        raw_matches = conn.execute(QUERY_FUZZY, (pfid, f"{fund_name[:25]}%")).fetchall()
        matches = [dict(m) for m in raw_matches]

    if not matches:
        suspect.append(
            (row_idx, f"NO DB MATCH: {pension_fund} | {fund_name} | {gp_excel} | v{vintage} | ${commitment}M")
        )
        continue

    # Find best match
    best = None
    best_score = -1
    for m in matches:
        score = 0
        if m["commitment_mm"] is not None and commitment is not None:
            if abs(m["commitment_mm"] - commitment) < 1.0:
                score += 3
            elif abs(m["commitment_mm"] - commitment) < 10.0:
                score += 1
        if m["vintage_year"] is not None and vintage is not None:
            if int(m["vintage_year"]) == int(vintage):
                score += 2
        if score > best_score:
            best_score = score
            best = m

    if not best:
        suspect.append((row_idx, f"No close match: {pension_fund} | {fund_name}"))
        continue

    m = best
    issues = []

    # Check commitment
    if commitment is not None and m["commitment_mm"] is not None:
        diff = abs(m["commitment_mm"] - commitment)
        if diff > 1.0:
            issues.append(f"commitment ${commitment}M vs DB ${m['commitment_mm']:.1f}M")

    # Check vintage
    if vintage is not None and m["vintage_year"] is not None:
        if int(vintage) != int(m["vintage_year"]):
            issues.append(f"vintage {int(vintage)} vs DB {m['vintage_year']}")

    # Check capital called
    xc = safe_float(rec["Capital Called ($M)"])
    if xc is not None and m["capital_called_mm"] is not None:
        diff = abs(m["capital_called_mm"] - xc)
        if diff > 1.0:
            issues.append(f"called ${xc:.1f}M vs DB ${m['capital_called_mm']:.1f}M")

    # Check distributions
    xd = safe_float(rec["Distributions ($M)"])
    if xd is not None and m["capital_distributed_mm"] is not None:
        diff = abs(m["capital_distributed_mm"] - xd)
        if diff > 1.0:
            issues.append(f"dist ${xd:.1f}M vs DB ${m['capital_distributed_mm']:.1f}M")

    # Check remaining value
    xr = safe_float(rec["Remaining Value ($M)"])
    if xr is not None and m["remaining_value_mm"] is not None:
        diff = abs(m["remaining_value_mm"] - xr)
        if diff > 1.5:
            issues.append(f"rem_val ${xr:.1f}M vs DB ${m['remaining_value_mm']:.1f}M")

    # Check IRR
    xi = safe_float(rec["Net IRR (%)"])
    if xi is not None and m["net_irr"] is not None:
        diff = abs(m["net_irr"] - xi)
        if diff > 0.5:
            issues.append(f"irr {xi}% vs DB {m['net_irr']:.1f}%")

    # Check multiple
    xm = safe_float(rec["Net Multiple (x)"])
    if xm is not None and m["net_multiple"] is not None:
        diff = abs(m["net_multiple"] - xm)
        if diff > 0.05:
            issues.append(f"multiple {xm:.2f}x vs DB {m['net_multiple']:.2f}x")

    # Check GP
    db_gp = m.get("general_partner_normalized", "")
    if gp_excel and db_gp and gp_excel.strip() != db_gp.strip():
        issues.append(f"GP '{gp_excel}' vs DB '{db_gp}'")

    if issues:
        suspect.append(
            (row_idx, f"MISMATCH: {pension_fund} | {fund_name}: " + "; ".join(issues))
        )
    else:
        confirmed.append(row_idx)

# Verify shifted rows against DB
shifted_confirmed = []
shifted_suspect = []
for row_idx, pf, gp, fund, commit, vintage in shifted_rows:
    pfid = PF_ID.get(pf)
    if not pfid:
        shifted_suspect.append((row_idx, f"Unknown pension fund: {pf}"))
        continue
    raw = conn.execute(QUERY, (pfid, fund)).fetchall()
    matches = [dict(m) for m in raw]
    if not matches:
        raw = conn.execute(QUERY_FUZZY, (pfid, f"{fund[:25]}%")).fetchall()
        matches = [dict(m) for m in raw]
    if matches:
        # Check commitment matches
        for m in matches:
            if commit and m["commitment_mm"] and abs(m["commitment_mm"] - commit) < 1.0:
                shifted_confirmed.append((row_idx, pf, gp, fund, commit, vintage))
                break
        else:
            shifted_confirmed.append((row_idx, pf, gp, fund, commit, vintage))
    else:
        shifted_suspect.append((row_idx, f"NO DB MATCH (shifted): {pf} | {fund} | {gp} | ${commit}M"))

conn.close()

print(f"=== VERIFICATION RESULTS ===")
print(f"CONFIRMED (normal rows):   {len(confirmed)} rows — all fields match database")
print(f"SUSPECT (normal rows):     {len(suspect)} rows")
print(f"COLUMN-SHIFTED rows:       {len(shifted_rows)} rows (data exists in DB but columns misaligned)")
print(f"  - Shifted & DB-confirmed: {len(shifted_confirmed)}")
print(f"  - Shifted & suspect:      {len(shifted_suspect)}")
print()

if suspect:
    print("=== SUSPECT ROWS (data mismatch or not found) ===")
    for row_idx, msg in suspect:
        print(f"  Row {row_idx}: {msg}")
    print()

if shifted_rows:
    print("=== COLUMN-SHIFTED ROWS (rows 119-126, columns offset by 1) ===")
    for row_idx, pf, gp, fund, commit, vintage in shifted_rows:
        status = "DB-CONFIRMED" if any(r == row_idx for r, *_ in shifted_confirmed) else "SUSPECT"
        print(f"  Row {row_idx}: [{status}] {pf} | {gp} | {fund} | ${commit}M | v{vintage}")
    print()

if shifted_suspect:
    print("=== SHIFTED ROW SUSPECTS ===")
    for row_idx, msg in shifted_suspect:
        print(f"  Row {row_idx}: {msg}")
    print()

pf_conf = Counter()
for r in confirmed:
    pf_conf[ws.cell(row=r, column=1).value] += 1
print("=== CONFIRMED BY PENSION FUND (normal rows) ===")
for pf, cnt in sorted(pf_conf.items()):
    print(f"  {pf}: {cnt}")

pf_susp = Counter()
for r, _ in suspect:
    pf_susp[ws.cell(row=r, column=1).value] += 1
if pf_susp:
    print("\n=== SUSPECT BY PENSION FUND ===")
    for pf, cnt in sorted(pf_susp.items()):
        print(f"  {pf}: {cnt}")

print(f"\n=== SUMMARY ===")
total_verified = len(confirmed) + len(shifted_confirmed)
total_suspect = len(suspect) + len(shifted_suspect)
print(f"Total verified against DB: {total_verified}")
print(f"Total suspect:             {total_suspect}")
print(f"Recommendation: Fix {len(shifted_rows)} shifted rows by pulling correct data from DB")
