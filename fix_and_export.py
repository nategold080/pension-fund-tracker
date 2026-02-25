"""
Jobs 1+2: Verify data, remove duplicates, fix shifted rows, fill N/D, export clean file.

Findings from verification:
- 111 rows: confirmed exactly against DB
- 6 NY Common rows: confirmed against 2025 report data (script initially matched 2024 report)
- 8 WSIB rows (119-126): column-shifted duplicates of rows 111-118 — REMOVE
- Total clean records: 117
"""

import sqlite3
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DB = "data/pension_tracker.db"
OUTPUT = "Dakota_Pension_Fund_Data_Sample_Feb19_.xlsx"

# ── Colors / styles ───────────────────────────────────────────────────────

NAVY = "1B2A4A"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F4F7"
MED_GRAY = "D9DCE3"
ACCENT_BLUE = "4472C4"
ALT_ROW = "EDF2F9"

thin_border = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)
header_border = Border(
    left=Side(style="thin", color=NAVY),
    right=Side(style="thin", color=NAVY),
    top=Side(style="thin", color=NAVY),
    bottom=Side(style="medium", color=NAVY),
)

# ── Load data ─────────────────────────────────────────────────────────────

CONSULTANT_MAP = {}
STATE_MAP = {}

MONTH_NAMES = {
    "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr",
    "05": "May", "06": "Jun", "07": "Jul", "08": "Aug",
    "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dec",
}


def fmt_date(d):
    """Convert '2025-03-31' to 'Mar 2025'."""
    s = str(d).strip()
    if len(s) >= 10 and s[4] == "-":
        return f"{MONTH_NAMES.get(s[5:7], s[5:7])} {s[:4]}"
    return s


def load_records():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row

    # Consultant map
    for r in conn.execute("""
        SELECT ce.pension_fund_id, cf.name, ce.role
        FROM consulting_engagements ce
        JOIN consulting_firms cf ON ce.consulting_firm_id = cf.id
    """).fetchall():
        pfid, name, role = r["pension_fund_id"], r["name"], r["role"]
        if pfid not in CONSULTANT_MAP or role == "general_investment_consultant":
            CONSULTANT_MAP[pfid] = name

    # State map
    for r in conn.execute("SELECT id, state FROM pension_funds").fetchall():
        STATE_MAP[r["id"]] = r["state"] or ""

    # Re-up lookup
    earliest = {}
    for r in conn.execute("""
        SELECT c.pension_fund_id, f.general_partner_normalized AS gp,
               MIN(f.vintage_year) AS ev
        FROM commitments c JOIN funds f ON c.fund_id = f.id
        WHERE f.general_partner_normalized IS NOT NULL
          AND f.general_partner_normalized != ''
        GROUP BY c.pension_fund_id, f.general_partner_normalized
    """).fetchall():
        earliest[(r["pension_fund_id"], r["gp"])] = r["ev"]

    # Main query: same selection logic as create_dakota_sample.py
    rows = conn.execute("""
        WITH gp_pension_counts AS (
            SELECT f.general_partner_normalized AS gp,
                   COUNT(DISTINCT c.pension_fund_id) AS pension_count
            FROM commitments c JOIN funds f ON c.fund_id = f.id
            WHERE f.general_partner_normalized IS NOT NULL
              AND f.general_partner_normalized != ''
            GROUP BY f.general_partner_normalized
        ),
        ranked AS (
            SELECT c.id, pf.name AS pension_fund, c.pension_fund_id,
                   f.general_partner_normalized AS gp, f.fund_name,
                   c.commitment_mm, c.capital_called_mm, c.capital_distributed_mm,
                   c.remaining_value_mm, c.net_irr, c.net_multiple,
                   f.asset_class, f.sub_strategy, f.vintage_year,
                   c.as_of_date, c.source_document, c.source_url,
                   c.extraction_method, c.extraction_confidence,
                   gpc.pension_count AS gp_cross_count,
                   ROW_NUMBER() OVER (
                       PARTITION BY c.pension_fund_id, f.general_partner_normalized
                       ORDER BY f.vintage_year DESC
                   ) AS rn
            FROM commitments c
            JOIN funds f ON c.fund_id = f.id
            JOIN pension_funds pf ON c.pension_fund_id = pf.id
            LEFT JOIN gp_pension_counts gpc ON gpc.gp = f.general_partner_normalized
            WHERE f.vintage_year >= 2015
              AND c.commitment_mm > 0
              AND f.general_partner_normalized IS NOT NULL
              AND f.general_partner_normalized != ''
        )
        SELECT * FROM ranked WHERE rn <= 2
        ORDER BY gp_cross_count DESC, gp, pension_fund, vintage_year DESC
    """).fetchall()

    non_pe = conn.execute("""
        SELECT c.id, pf.name AS pension_fund, c.pension_fund_id,
               f.general_partner_normalized AS gp, f.fund_name,
               c.commitment_mm, c.capital_called_mm, c.capital_distributed_mm,
               c.remaining_value_mm, c.net_irr, c.net_multiple,
               f.asset_class, f.sub_strategy, f.vintage_year,
               c.as_of_date, c.source_document, c.source_url,
               c.extraction_method, c.extraction_confidence,
               1 AS gp_cross_count, 1 AS rn
        FROM commitments c
        JOIN funds f ON c.fund_id = f.id
        JOIN pension_funds pf ON c.pension_fund_id = pf.id
        WHERE f.asset_class != 'Private Equity'
          AND f.vintage_year >= 2015
          AND c.commitment_mm > 0
          AND f.general_partner_normalized IS NOT NULL
          AND f.general_partner_normalized != ''
        ORDER BY f.vintage_year DESC
    """).fetchall()

    conn.close()

    def enrich(row):
        rec = dict(row)
        key = (rec["pension_fund_id"], rec["gp"])
        ev = earliest.get(key)
        rec["commitment_type"] = "Re-Up" if (ev and ev < (rec["vintage_year"] or 0)) else "New Relationship"
        rec["consultant"] = CONSULTANT_MAP.get(rec["pension_fund_id"], "")
        rec["state"] = STATE_MAP.get(rec["pension_fund_id"], "")
        # Fill TA sub-strategy
        if rec["gp"] == "TA" and not rec["sub_strategy"]:
            rec["sub_strategy"] = "Growth Equity"
        return rec

    selected = []
    seen = set()

    for row in rows:
        if len(selected) >= 110:
            break
        if row["id"] not in seen:
            selected.append(enrich(row))
            seen.add(row["id"])

    for row in non_pe:
        if len(selected) >= 125:
            break
        if row["id"] not in seen:
            selected.append(enrich(row))
            seen.add(row["id"])

    # Dedup: keep most recent as_of_date per (pension_fund_id, fund_name, vintage_year)
    best = {}
    for rec in selected:
        k = (rec["pension_fund_id"], rec["fund_name"], rec["vintage_year"])
        if k not in best or (rec["as_of_date"] or "") > (best[k]["as_of_date"] or ""):
            best[k] = rec

    deduped = list(best.values())

    # Format dates
    for rec in deduped:
        rec["as_of_date"] = fmt_date(rec.get("as_of_date", ""))

    return deduped


# ── Column definitions ────────────────────────────────────────────────────

COLUMNS = [
    ("Pension Fund",         22, "pension_fund",          "text"),
    ("State",                8,  "state",                 "center"),
    ("GP Name",              28, "gp",                    "text"),
    ("Fund Name",            42, "fund_name",             "text"),
    ("Asset Class",          18, "asset_class",           "text"),
    ("Sub-Strategy",         24, "sub_strategy",          "text"),
    ("Vintage Year",         14, "vintage_year",          "center"),
    ("Commitment ($M)",      18, "commitment_mm",         "currency"),
    ("Capital Called ($M)",  20, "capital_called_mm",     "currency"),
    ("Distributions ($M)",   20, "capital_distributed_mm","currency"),
    ("Remaining Value ($M)", 22, "remaining_value_mm",    "currency"),
    ("Net IRR (%)",          14, "net_irr",               "pct"),
    ("Net Multiple (x)",     16, "net_multiple",          "multiple"),
    ("Source Report Date",   16, "as_of_date",            "center"),
    ("Commitment Type",      18, "commitment_type",       "committype"),
    ("Investment Consultant",24, "consultant",            "text"),
    ("Source Document",      36, "source_document",       "text"),
    ("Source URL",           50, "source_url",            "url"),
    ("Extraction Method",    22, "extraction_method",     "center"),
    ("Confidence",           12, "extraction_confidence", "confidence"),
]

# Fields where blank means "Not Disclosed by source"
ND_FIELDS = {
    "sub_strategy", "capital_called_mm", "capital_distributed_mm",
    "remaining_value_mm", "net_irr", "net_multiple",
}


def build_excel(records):
    records.sort(key=lambda r: (r["pension_fund"], -(r["vintage_year"] or 0)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.sheet_properties.tabColor = ACCENT_BLUE

    for i, (_, width, _, _) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Header
    hdr_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    hdr_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, (header, _, _, _) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=header)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = header_border

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # Style presets
    even_fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
    odd_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    dfont = Font(name="Calibri", size=10, color="333333")
    nd_font = Font(name="Calibri", size=10, color="999999")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    url_font = Font(name="Calibri", size=9, color="666666")

    for ri, rec in enumerate(records, 2):
        fill = even_fill if ri % 2 == 0 else odd_fill

        for ci, (_, _, key, fmt) in enumerate(COLUMNS, 1):
            val = rec.get(key)
            c = ws.cell(row=ri, column=ci)
            c.fill = fill
            c.border = thin_border

            # Determine if this is N/D
            is_nd = (val is None or val == "") and key in ND_FIELDS

            if is_nd:
                c.value = "N/D"
                c.font = nd_font
                c.alignment = center
            elif fmt == "currency":
                c.value = val if val not in (None, "") else None
                c.number_format = '#,##0.0'
                c.alignment = right
                c.font = dfont
            elif fmt == "pct":
                c.value = val if val not in (None, "") else None
                c.number_format = '0.0'
                c.alignment = right
                c.font = dfont
            elif fmt == "multiple":
                c.value = val if val not in (None, "") else None
                c.number_format = '0.00"x"'
                c.alignment = right
                c.font = dfont
            elif fmt == "confidence":
                c.value = val if val not in (None, "") else None
                c.number_format = '0.00'
                c.alignment = center
                c.font = dfont
            elif fmt == "center":
                c.value = val if val not in (None, "") else None
                c.alignment = center
                c.font = dfont
            elif fmt == "committype":
                c.value = val
                c.alignment = center
                if val == "Re-Up":
                    c.font = Font(name="Calibri", size=10, color=ACCENT_BLUE, bold=True)
                else:
                    c.font = Font(name="Calibri", size=10, color="666666")
            elif fmt == "url":
                c.value = val if val not in (None, "") else None
                c.alignment = left
                c.font = url_font
            else:
                c.value = val if val not in (None, "") else None
                c.alignment = left
                c.font = dfont

        ws.row_dimensions[ri].height = 20

    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{len(records) + 1}"

    wb.save(OUTPUT)
    return len(records)


def main():
    print("Loading verified records from database...")
    records = load_records()
    print(f"Records after dedup: {len(records)}")

    # Count stats
    pf_counts = defaultdict(int)
    gps = set()
    funds = set()
    for r in records:
        pf_counts[r["pension_fund"]] += 1
        gps.add(r["gp"])
        funds.add(r["fund_name"])

    for pf, cnt in sorted(pf_counts.items()):
        print(f"  {pf}: {cnt}")
    print(f"  Unique GPs: {len(gps)}, Unique Funds: {len(funds)}")

    # Count blanks that will become N/D
    nd_count = 0
    for rec in records:
        for key in ND_FIELDS:
            if rec.get(key) is None or rec.get(key) == "":
                nd_count += 1
    print(f"\nBlanks to fill with N/D: {nd_count}")

    print("\nBuilding Excel...")
    n = build_excel(records)
    print(f"Saved {OUTPUT} with {n} records")


if __name__ == "__main__":
    main()
