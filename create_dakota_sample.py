"""Generate Dakota Marketplace sample data Excel file."""

import sqlite3
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

DB_PATH = "data/pension_tracker.db"
OUTPUT_PATH = "Dakota_Pension_Fund_Data_Sample.xlsx"

# ── Consulting engagement lookup ──────────────────────────────────────────

CONSULTANT_MAP: dict[str, str] = {}


def load_consultant_map(conn: sqlite3.Connection) -> None:
    rows = conn.execute("""
        SELECT ce.pension_fund_id, cf.name, ce.role
        FROM consulting_engagements ce
        JOIN consulting_firms cf ON ce.consulting_firm_id = cf.id
        ORDER BY ce.pension_fund_id, ce.role
    """).fetchall()
    for pfid, name, role in rows:
        if pfid not in CONSULTANT_MAP:
            CONSULTANT_MAP[pfid] = name
        if role == "general_investment_consultant":
            CONSULTANT_MAP[pfid] = name


# ── State lookup ──────────────────────────────────────────────────────────

STATE_MAP: dict[str, str] = {}


def load_state_map(conn: sqlite3.Connection) -> None:
    rows = conn.execute("SELECT id, state FROM pension_funds").fetchall()
    for pfid, state in rows:
        STATE_MAP[pfid] = state or ""


# ── Query sample records ──────────────────────────────────────────────────

SAMPLE_QUERY = """
WITH gp_pension_counts AS (
    SELECT f.general_partner_normalized AS gp,
           COUNT(DISTINCT c.pension_fund_id) AS pension_count
    FROM commitments c
    JOIN funds f ON c.fund_id = f.id
    WHERE f.general_partner_normalized IS NOT NULL
      AND f.general_partner_normalized != ''
    GROUP BY f.general_partner_normalized
),
ranked AS (
    SELECT c.id,
           pf.name AS pension_fund,
           c.pension_fund_id,
           f.general_partner_normalized AS gp,
           f.fund_name,
           c.commitment_mm,
           c.capital_called_mm,
           c.capital_distributed_mm,
           c.remaining_value_mm,
           c.net_irr,
           c.net_multiple,
           f.asset_class,
           f.sub_strategy,
           f.vintage_year,
           c.as_of_date,
           c.source_document,
           c.source_url,
           c.extraction_method,
           c.extraction_confidence,
           gpc.pension_count AS gp_cross_count,
           ROW_NUMBER() OVER (
               PARTITION BY c.pension_fund_id, f.general_partner_normalized
               ORDER BY f.vintage_year DESC
           ) AS rn
    FROM commitments c
    JOIN funds f ON c.fund_id = f.id
    JOIN pension_funds pf ON c.pension_fund_id = pf.id
    LEFT JOIN gp_pension_counts gpc
        ON gpc.gp = f.general_partner_normalized
    WHERE f.vintage_year >= 2015
      AND c.commitment_mm > 0
      AND f.general_partner_normalized IS NOT NULL
      AND f.general_partner_normalized != ''
)
SELECT *
FROM ranked
WHERE rn <= 2
ORDER BY gp_cross_count DESC, gp, pension_fund, vintage_year DESC
"""

NON_PE_QUERY = """
SELECT c.id,
       pf.name AS pension_fund,
       c.pension_fund_id,
       f.general_partner_normalized AS gp,
       f.fund_name,
       c.commitment_mm,
       c.capital_called_mm,
       c.capital_distributed_mm,
       c.remaining_value_mm,
       c.net_irr,
       c.net_multiple,
       f.asset_class,
       f.sub_strategy,
       f.vintage_year,
       c.as_of_date,
       c.source_document,
       c.source_url,
       c.extraction_method,
       c.extraction_confidence,
       1 AS gp_cross_count,
       1 AS rn
FROM commitments c
JOIN funds f ON c.fund_id = f.id
JOIN pension_funds pf ON c.pension_fund_id = pf.id
WHERE f.asset_class != 'Private Equity'
  AND f.vintage_year >= 2015
  AND c.commitment_mm > 0
  AND f.general_partner_normalized IS NOT NULL
  AND f.general_partner_normalized != ''
ORDER BY f.vintage_year DESC
"""

REUP_QUERY = """
SELECT c.pension_fund_id, f.general_partner_normalized AS gp,
       MIN(f.vintage_year) AS earliest_vintage
FROM commitments c
JOIN funds f ON c.fund_id = f.id
WHERE f.general_partner_normalized IS NOT NULL
  AND f.general_partner_normalized != ''
GROUP BY c.pension_fund_id, f.general_partner_normalized
"""


def load_data():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    load_consultant_map(conn)
    load_state_map(conn)

    reup_rows = conn.execute(REUP_QUERY).fetchall()
    earliest_vintage: dict[tuple[str, str], int] = {}
    for r in reup_rows:
        earliest_vintage[(r["pension_fund_id"], r["gp"])] = r["earliest_vintage"]

    rows = conn.execute(SAMPLE_QUERY).fetchall()
    non_pe_rows = conn.execute(NON_PE_QUERY).fetchall()

    def enrich(row) -> dict:
        rec = dict(row)
        key = (rec["pension_fund_id"], rec["gp"])
        ev = earliest_vintage.get(key)
        if ev and ev < (rec["vintage_year"] or 0):
            rec["commitment_type"] = "Re-Up"
        else:
            rec["commitment_type"] = "New Relationship"
        rec["consultant"] = CONSULTANT_MAP.get(rec["pension_fund_id"], "")
        rec["state"] = STATE_MAP.get(rec["pension_fund_id"], "")
        return rec

    selected: list[dict] = []
    seen_ids: set[str] = set()
    pension_counts: dict[str, int] = defaultdict(int)
    PE_TARGET = 110
    TOTAL_TARGET = 125

    for row in rows:
        if len(selected) >= PE_TARGET:
            break
        rid = row["id"]
        if rid in seen_ids:
            continue
        rec = enrich(row)
        selected.append(rec)
        seen_ids.add(rid)
        pension_counts[rec["pension_fund"]] += 1

    for row in non_pe_rows:
        if len(selected) >= TOTAL_TARGET:
            break
        rid = row["id"]
        if rid in seen_ids:
            continue
        rec = enrich(row)
        selected.append(rec)
        seen_ids.add(rid)
        pension_counts[rec["pension_fund"]] += 1

    conn.close()

    # ── Post-processing fixes ──

    # 1. Fill missing sub-strategy for TA Associates → "Growth Equity"
    for rec in selected:
        if rec["gp"] == "TA" and not rec["sub_strategy"]:
            rec["sub_strategy"] = "Growth Equity"

    # 2. Deduplicate NY Common: same fund appearing with different as_of_date.
    #    Keep only the most recent report date per (pension_fund, fund_name, vintage_year).
    dedup_key = lambda r: (r["pension_fund_id"], r["fund_name"], r["vintage_year"])
    best: dict[tuple, dict] = {}
    for rec in selected:
        k = dedup_key(rec)
        if k not in best or (rec["as_of_date"] or "") > (best[k]["as_of_date"] or ""):
            best[k] = rec
    deduped = list(best.values())

    removed = len(selected) - len(deduped)
    if removed:
        print(f"  Deduplicated: removed {removed} duplicate rows")

    # Recount per pension fund
    pension_counts = defaultdict(int)
    for rec in deduped:
        pension_counts[rec["pension_fund"]] += 1

    # 3. Clean date formatting: "2025-03-31" → "Mar 2025"
    month_names = {
        "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr",
        "05": "May", "06": "Jun", "07": "Jul", "08": "Aug",
        "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dec",
    }
    for rec in deduped:
        d = rec.get("as_of_date") or ""
        d = str(d).strip()
        # Handle "2025-03-31 00:00:00" or "2025-03-31"
        if len(d) >= 10 and d[4] == "-":
            mm = d[5:7]
            yyyy = d[:4]
            rec["as_of_date"] = f"{month_names.get(mm, mm)} {yyyy}"

    return deduped, pension_counts


# ── Statistics ────────────────────────────────────────────────────────────

def compute_stats(records: list[dict]) -> dict:
    pension_funds = set()
    gps = set()
    funds = set()
    asset_classes = set()
    vintages = set()
    total_commitment = 0.0

    for r in records:
        pension_funds.add(r["pension_fund"])
        gps.add(r["gp"])
        funds.add(r["fund_name"])
        if r["asset_class"]:
            asset_classes.add(r["asset_class"])
        if r["vintage_year"]:
            vintages.add(r["vintage_year"])
        total_commitment += r["commitment_mm"] or 0

    return {
        "pension_systems": len(pension_funds),
        "pension_names": sorted(pension_funds),
        "total_records": len(records),
        "unique_gps": len(gps),
        "unique_funds": len(funds),
        "asset_classes": sorted(asset_classes),
        "vintage_min": min(vintages) if vintages else None,
        "vintage_max": max(vintages) if vintages else None,
        "total_commitment_mm": total_commitment,
    }


def compute_full_db_stats() -> dict:
    conn = sqlite3.connect(DB_PATH)
    total_records = conn.execute("SELECT COUNT(*) FROM commitments").fetchone()[0]
    total_funds = conn.execute("SELECT COUNT(*) FROM funds").fetchone()[0]
    total_gps = conn.execute(
        "SELECT COUNT(DISTINCT general_partner_normalized) FROM funds "
        "WHERE general_partner_normalized IS NOT NULL AND general_partner_normalized != ''"
    ).fetchone()[0]
    pension_systems = conn.execute(
        "SELECT COUNT(DISTINCT pension_fund_id) FROM commitments"
    ).fetchone()[0]
    pension_names = [
        r[0] for r in conn.execute(
            "SELECT DISTINCT pf.name FROM commitments c "
            "JOIN pension_funds pf ON c.pension_fund_id = pf.id ORDER BY pf.name"
        ).fetchall()
    ]
    vintage_range = conn.execute(
        "SELECT MIN(f.vintage_year), MAX(f.vintage_year) "
        "FROM commitments c JOIN funds f ON c.fund_id = f.id "
        "WHERE f.vintage_year IS NOT NULL"
    ).fetchone()
    cross_linked = conn.execute("""
        SELECT COUNT(*) FROM (
            SELECT f.id FROM commitments c JOIN funds f ON c.fund_id = f.id
            GROUP BY f.id HAVING COUNT(DISTINCT c.pension_fund_id) >= 2
        )
    """).fetchone()[0]
    total_commitment = conn.execute(
        "SELECT SUM(commitment_mm) FROM commitments"
    ).fetchone()[0]
    conn.close()
    return {
        "total_records": total_records,
        "total_funds": total_funds,
        "total_gps": total_gps,
        "pension_systems": pension_systems,
        "pension_names": pension_names,
        "vintage_min": vintage_range[0],
        "vintage_max": vintage_range[1],
        "cross_linked_funds": cross_linked,
        "total_commitment_mm": total_commitment or 0,
    }


# ── Excel styling constants ──────────────────────────────────────────────

NAVY = "1B2A4A"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F4F7"
MED_GRAY = "D9DCE3"
ACCENT_BLUE = "4472C4"
ALT_ROW = "EDF2F9"

thin_border = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)

header_border = Border(
    left=Side(style="thin", color=NAVY),
    right=Side(style="thin", color=NAVY),
    top=Side(style="thin", color=NAVY),
    bottom=Side(style="medium", color=NAVY),
)


# ── Overview sheet ────────────────────────────────────────────────────────

def create_overview_sheet(wb: Workbook, stats: dict, db_stats: dict) -> None:
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_properties.tabColor = NAVY

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 65
    ws.column_dimensions["D"].width = 3

    row = 2

    # Title
    ws.merge_cells("B2:C2")
    c = ws.cell(row=row, column=2)
    c.value = "Pension Fund Commitment Intelligence"
    c.font = Font(name="Calibri", size=22, bold=True, color=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 38
    row += 1

    ws.merge_cells("B3:C3")
    c = ws.cell(row=row, column=2)
    c.value = "Sample Dataset"
    c.font = Font(name="Calibri", size=16, color=ACCENT_BLUE)
    c.alignment = Alignment(horizontal="left")
    ws.row_dimensions[row].height = 28
    row += 2

    # Contact info
    info_items = [
        ("Prepared by:", "Nathan Goldberg"),
        ("Email:", "nathanmauricegoldberg@gmail.com"),
        ("Date:", "February 19, 2026"),
        ("Prepared for:", "Rob Robertson, President — Dakota Marketplace"),
    ]
    for label, value in info_items:
        ws.cell(row=row, column=2, value=label).font = Font(
            name="Calibri", size=11, bold=True, color=NAVY
        )
        ws.cell(row=row, column=3, value=value).font = Font(
            name="Calibri", size=11, color="333333"
        )
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1

    # Divider
    for col in (2, 3):
        ws.cell(row=row, column=col).border = Border(
            bottom=Side(style="medium", color=ACCENT_BLUE)
        )
    row += 2

    # Description
    ws.merge_cells(f"B{row}:C{row}")
    ws.cell(row=row, column=2, value="About This Dataset").font = Font(
        name="Calibri", size=14, bold=True, color=NAVY
    )
    ws.row_dimensions[row].height = 24
    row += 1

    desc_text = (
        "This dataset contains structured pension fund commitment data extracted from "
        "public board meeting minutes, quarterly investment reports, and statutory "
        f"disclosure documents. It covers {db_stats['pension_systems']} major U.S. public "
        f"pension systems with {db_stats['total_records']:,} entity-resolved commitment "
        f"records spanning {db_stats['total_gps']:,} unique general partners and "
        f"{db_stats['total_funds']:,} funds. "
        "All data is deterministically extracted from primary sources (HTML tables and "
        "PDF filings) — no LLM-generated content — with full provenance tracking to "
        "the original document, page, and extraction method."
    )
    ws.merge_cells(f"B{row}:C{row}")
    c = ws.cell(row=row, column=2, value=desc_text)
    c.font = Font(name="Calibri", size=11, color="333333")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 80
    row += 2

    # Full Database Summary
    ws.merge_cells(f"B{row}:C{row}")
    ws.cell(row=row, column=2, value="Full Database Summary").font = Font(
        name="Calibri", size=14, bold=True, color=NAVY
    )
    ws.row_dimensions[row].height = 24
    row += 1

    def write_stat_table(items):
        nonlocal row
        for label, value in items:
            for col, val, bold in [(2, label, True), (3, value, False)]:
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name="Calibri", size=11, bold=bold, color=NAVY if bold else "333333")
                c.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
                c.border = thin_border
            ws.row_dimensions[row].height = 22
            row += 1

    write_stat_table([
        ("Pension Systems Covered", f"{db_stats['pension_systems']} ({', '.join(db_stats['pension_names'])})"),
        ("Total Commitment Records", f"{db_stats['total_records']:,}"),
        ("Unique Funds", f"{db_stats['total_funds']:,}"),
        ("Unique General Partners", f"{db_stats['total_gps']:,}"),
        ("Cross-Linked Funds (2+ systems)", f"{db_stats['cross_linked_funds']:,}"),
        ("Total Commitments Tracked", f"${db_stats['total_commitment_mm']:,.0f}M"),
        ("Vintage Year Range", f"{db_stats['vintage_min']} – {db_stats['vintage_max']}"),
    ])

    row += 1
    for col in (2, 3):
        ws.cell(row=row, column=col).border = Border(
            bottom=Side(style="medium", color=ACCENT_BLUE)
        )
    row += 2

    # Sample Data Summary
    ws.merge_cells(f"B{row}:C{row}")
    ws.cell(row=row, column=2, value="Sample Data Summary (Sheet 2)").font = Font(
        name="Calibri", size=14, bold=True, color=NAVY
    )
    ws.row_dimensions[row].height = 24
    row += 1

    write_stat_table([
        ("Records in Sample", f"{stats['total_records']}"),
        ("Pension Systems", f"{stats['pension_systems']} ({', '.join(stats['pension_names'])})"),
        ("Unique GPs in Sample", f"{stats['unique_gps']}"),
        ("Unique Funds in Sample", f"{stats['unique_funds']}"),
        ("Vintage Year Range", f"{stats['vintage_min']} – {stats['vintage_max']}"),
        ("Total Commitment Value", f"${stats['total_commitment_mm']:,.0f}M"),
        ("Asset Classes", ", ".join(stats["asset_classes"])),
    ])

    row += 2

    # Field Definitions
    ws.merge_cells(f"B{row}:C{row}")
    ws.cell(row=row, column=2, value="Field Definitions").font = Font(
        name="Calibri", size=14, bold=True, color=NAVY
    )
    ws.row_dimensions[row].height = 24
    row += 1

    field_defs = [
        ("Pension Fund", "The public pension system (limited partner / LP) that made the commitment."),
        ("State", "U.S. state where the pension system is domiciled."),
        ("GP Name", "The general partner (investment manager) managing the fund."),
        ("Fund Name", "The specific investment vehicle to which the commitment was made."),
        ("Asset Class", "High-level asset category: Private Equity, Private Credit, Real Assets."),
        ("Sub-Strategy", "Investment strategy within asset class: Buyout, Growth Equity, Venture Capital, Credit, Infrastructure, etc."),
        ("Vintage Year", "The year the fund began making investments (fund formation year)."),
        ("Commitment ($M)", "The total capital commitment in millions of U.S. dollars."),
        ("Capital Called ($M)", "Capital drawn down by the GP from the commitment, in millions."),
        ("Distributions ($M)", "Capital returned to the LP (pension fund), in millions."),
        ("Remaining Value ($M)", "Current net asset value (NAV) of the pension fund's position, in millions."),
        ("Net IRR (%)", "Net internal rate of return after fees. Blank if not disclosed by the source."),
        ("Net Multiple (x)", "Total value to paid-in capital ratio (TVPI). Values > 1.0x indicate positive returns."),
        ("Source Report Date", "The reporting period end date of the source document from which data was extracted."),
        ("Commitment Type", "\"Re-Up\" if the pension fund previously committed to an earlier fund from the same GP; otherwise \"New Relationship.\""),
        ("Investment Consultant", "The pension fund's primary investment consultant/advisor (from public filings)."),
        ("Source Document", "The public document from which the data was extracted."),
        ("Source URL", "Direct URL to the source document or data page."),
        ("Extraction Method", "How the data was extracted: deterministic_html (web tables) or deterministic_pdf (document parsing)."),
        ("Confidence", "Extraction confidence score (0.0–1.0). 1.0 = fully deterministic; lower scores indicate heuristic-assisted parsing."),
    ]

    for col_idx, header in enumerate(["Field", "Definition"], start=2):
        c = ws.cell(row=row, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        c.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        c.border = header_border
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24
    row += 1

    for i, (field, defn) in enumerate(field_defs):
        fill_color = LIGHT_GRAY if i % 2 == 0 else WHITE
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        c = ws.cell(row=row, column=2, value=field)
        c.font = Font(name="Calibri", size=10, bold=True, color="333333")
        c.fill = fill
        c.border = thin_border
        c = ws.cell(row=row, column=3, value=defn)
        c.font = Font(name="Calibri", size=10, color="333333")
        c.fill = fill
        c.border = thin_border
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1

    row += 2

    # Methodology
    ws.merge_cells(f"B{row}:C{row}")
    ws.cell(row=row, column=2, value="Methodology").font = Font(
        name="Calibri", size=14, bold=True, color=NAVY
    )
    ws.row_dimensions[row].height = 24
    row += 1

    methodology = (
        "Data is extracted using a deterministic, source-aware pipeline. Each pension "
        "system has a dedicated adapter that parses its specific disclosure format — "
        "HTML tables for web-published data, word-level PDF extraction for document-based "
        "disclosures. No LLM or AI is used for data extraction; all parsing is rule-based "
        "with full provenance tracking to source document, page, and extraction method. "
        "Fund names are entity-resolved across pension systems using a master registry "
        "with canonical GP names, vintage years, and alias mappings, enabling cross-system "
        "analysis of the same underlying fund."
    )
    ws.merge_cells(f"B{row}:C{row}")
    c = ws.cell(row=row, column=2, value=methodology)
    c.font = Font(name="Calibri", size=11, color="333333")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 85
    row += 2

    # Footer
    ws.merge_cells(f"B{row}:C{row}")
    c = ws.cell(row=row, column=2)
    c.value = "CONFIDENTIAL — Prepared for Dakota Marketplace evaluation purposes only."
    c.font = Font(name="Calibri", size=9, italic=True, color="999999")
    c.alignment = Alignment(horizontal="center")


# ── Data sheet ────────────────────────────────────────────────────────────

# Column spec: (header, width, data_key, fmt_type)
# fmt_type: "text", "currency", "pct", "multiple", "center", "url", "confidence"
DATA_COLUMNS = [
    ("Pension Fund",        22, "pension_fund",          "text"),
    ("State",               8,  "state",                 "center"),
    ("GP Name",             28, "gp",                    "text"),
    ("Fund Name",           42, "fund_name",             "text"),
    ("Asset Class",         18, "asset_class",           "text"),
    ("Sub-Strategy",        24, "sub_strategy",          "text"),
    ("Vintage Year",        14, "vintage_year",          "center"),
    ("Commitment ($M)",     18, "commitment_mm",         "currency"),
    ("Capital Called ($M)", 20, "capital_called_mm",     "currency"),
    ("Distributions ($M)",  20, "capital_distributed_mm","currency"),
    ("Remaining Value ($M)",22, "remaining_value_mm",    "currency"),
    ("Net IRR (%)",         14, "net_irr",               "pct"),
    ("Net Multiple (x)",    16, "net_multiple",          "multiple"),
    ("Source Report Date",  16, "as_of_date",            "center"),
    ("Commitment Type",     18, "commitment_type",       "committype"),
    ("Investment Consultant",24,"consultant",            "text"),
    ("Source Document",      36, "source_document",      "text"),
    ("Source URL",           50, "source_url",            "url"),
    ("Extraction Method",    22, "extraction_method",     "center"),
    ("Confidence",           12, "extraction_confidence", "confidence"),
]


def create_data_sheet(wb: Workbook, records: list[dict]) -> None:
    ws = wb.create_sheet("Sample Data")
    ws.sheet_properties.tabColor = ACCENT_BLUE

    # Sort: pension fund (asc), then vintage year (desc)
    records.sort(key=lambda r: (r["pension_fund"], -(r["vintage_year"] or 0)))

    # Column widths
    for i, (_, width, _, _) in enumerate(DATA_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Header row
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (header, _, _, _) in enumerate(DATA_COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = header_border

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # Style presets
    even_fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
    odd_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    data_font = Font(name="Calibri", size=10, color="333333")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    url_font = Font(name="Calibri", size=9, color="666666")

    for row_idx, rec in enumerate(records, start=2):
        fill = even_fill if row_idx % 2 == 0 else odd_fill

        for col_idx, (_, _, key, fmt) in enumerate(DATA_COLUMNS, start=1):
            val = rec.get(key)
            if val is None:
                val = ""

            c = ws.cell(row=row_idx, column=col_idx)
            c.fill = fill
            c.border = thin_border

            if fmt == "currency":
                c.value = val if val != "" else None
                c.number_format = '#,##0.0'
                c.alignment = right_align
                c.font = data_font
            elif fmt == "pct":
                if val != "" and val is not None:
                    c.value = val
                    c.number_format = '0.0'
                else:
                    c.value = None
                c.alignment = right_align
                c.font = data_font
            elif fmt == "multiple":
                if val != "" and val is not None:
                    c.value = val
                    c.number_format = '0.00"x"'
                else:
                    c.value = None
                c.alignment = right_align
                c.font = data_font
            elif fmt == "confidence":
                if val != "" and val is not None:
                    c.value = val
                    c.number_format = '0.00'
                else:
                    c.value = None
                c.alignment = center_align
                c.font = data_font
            elif fmt == "center":
                c.value = val if val != "" else None
                c.alignment = center_align
                c.font = data_font
            elif fmt == "committype":
                c.value = val
                c.alignment = center_align
                if val == "Re-Up":
                    c.font = Font(name="Calibri", size=10, color=ACCENT_BLUE, bold=True)
                else:
                    c.font = Font(name="Calibri", size=10, color="666666")
            elif fmt == "url":
                c.value = val if val != "" else None
                c.alignment = left_align
                c.font = url_font
            else:  # text
                c.value = val if val != "" else None
                c.alignment = left_align
                c.font = data_font

        ws.row_dimensions[row_idx].height = 20

    # Auto-filter
    last_col = get_column_letter(len(DATA_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{len(records) + 1}"


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    print("Loading data from database...")
    records, pension_counts = load_data()

    print(f"Selected {len(records)} records")
    for pf, cnt in sorted(pension_counts.items()):
        print(f"  {pf}: {cnt} records")

    stats = compute_stats(records)
    db_stats = compute_full_db_stats()

    print(f"\nFull database: {db_stats['total_records']:,} records, "
          f"{db_stats['total_gps']:,} GPs, {db_stats['total_funds']:,} funds")

    wb = Workbook()

    print("Creating Overview sheet...")
    create_overview_sheet(wb, stats, db_stats)

    print("Creating Sample Data sheet...")
    create_data_sheet(wb, records)

    wb.save(OUTPUT_PATH)
    print(f"\nSaved to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
